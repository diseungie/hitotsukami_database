# プログラム0　「ライブラリの設定（下準備）」
from bs4 import BeautifulSoup
import urllib.request as req
from selenium import webdriver
import time
import pandas as pd
from datetime import datetime
import openpyxl as px
from openpyxl.styles import PatternFill
import requests
import numpy as np
from selenium.webdriver.support.ui import Select

risyu_list = []


# プログラム1-1　「GoogleChromeを起動」
browser = webdriver.Chrome(executable_path = 'C:\\Users\\Sota2020\\Desktop\\MyPandas\\chromedriver.exe')
browser.implicitly_wait(1)


# プログラム1-2　「学外用シラバス検索サイトへアクセス」
url_syllabus = "https://syllabus.cels.hit-u.ac.jp/"
browser.get(url_syllabus)
time.sleep(1)


# プログラム2　「月～金までの1～6限の授業情報を取得」

for yobi in range(1, 6):
    for jigen in range(1, 7):
# プログラム2-1　「ドロップダウンメニュー選択（年度）」
        nendo_element = browser.find_element_by_name('nendo')
        nendo_select_element = Select(nendo_element)
        nendo_select_element.select_by_value('2021')

# プログラム2-2　「ドロップダウンメニュー選択（曜日）」
        yobi_element = browser.find_element_by_name('yobi')
        yobi_select_element = Select(yobi_element)
        yobi_select_element.select_by_value(str(yobi))

# プログラム2-3　「ドロップダウンメニュー選択（時限）」
        jigen_element = browser.find_element_by_name('jigen')
        jigen_select_element = Select(jigen_element)
        jigen_select_element.select_by_value(str(jigen))

# プログラム2-4　「ドロップダウンメニュー選択（検索結果表示件数）」
        _displayCount_element = browser.find_element_by_name('_displayCount')
        _displayCount_select_element = Select(_displayCount_element)
        _displayCount_select_element.select_by_value('200')

# プログラム2-5　「入力した検索条件で検索開始」
        browser_from = browser.find_element_by_name('_eventId_search')
        #time.sleep(1)
        browser_from.click()

# プログラム2-6　「授業科目一覧ページのURLを取得」
        cur_url = browser.current_url


# プログラム2-7　「授業科目一覧ページ（ブラウザで表示中）の全ソースを取得」
        cur_source = browser.page_source


# プログラム2-8　「HTML解析と必要情報取得（と整理）」
        soup = BeautifulSoup(cur_source, 'lxml')
        TD = soup.find_all('td')

        for j in TD:
            risyu_list.append(j.text)
        

# プログラム2-9　「検索画面へ戻る」
        browser.back()    
    

# プログラム3　「曜日：他 の授業情報を取得」   
    
# プログラム3-1　「ドロップダウンメニュー選択（年度）」
nendo_element = browser.find_element_by_name('nendo')
nendo_select_element = Select(nendo_element)
nendo_select_element.select_by_value('2021')

# プログラム3-2　「ドロップダウンメニュー選択（曜日）」
yobi_element = browser.find_element_by_name('yobi')
yobi_select_element = Select(yobi_element)
yobi_select_element.select_by_value('9')

# プログラム3-3　「ドロップダウンメニュー選択（時限）」
jigen_element = browser.find_element_by_name('jigen')
jigen_select_element = Select(jigen_element)
jigen_select_element.select_by_visible_text('指示なし')

# プログラム3-4　「ドロップダウンメニュー選択（検索結果表示件数）」
_displayCount_element = browser.find_element_by_name('_displayCount')
_displayCount_select_element = Select(_displayCount_element)
_displayCount_select_element.select_by_value('200')

# プログラム3-5　「入力した検索条件で検索開始」
browser_from = browser.find_element_by_name('_eventId_search')
browser_from.click()

# プログラム3-6-1　「授業科目一覧ページのURLを取得」
cur_url = browser.current_url

# プログラム3-6-2　「授業科目一覧ページ（ブラウザで表示中）の全ソースを取得」
cur_source = browser.page_source

# プログラム3-6-3　「HTML解析と必要情報取得（と整理）」
soup = BeautifulSoup(cur_source, 'lxml')
TD = soup.find_all('td')

for j in TD:
    risyu_list.append(j.text)

# プログラム3-7-0　「授業科目一覧ページ2へ移動」        
page_second = browser.find_element_by_link_text('2')
page_second.click()

# プログラム3-7-1　「授業科目一覧ページのURLを取得」
cur_url = browser.current_url

# プログラム3-7-2　「授業科目一覧ページ（ブラウザで表示中）の全ソースを取得」
cur_source = browser.page_source

# プログラム3-7-3　「HTML解析と必要情報取得（と整理）」
soup = BeautifulSoup(cur_source, 'lxml')
TD = soup.find_all('td')

for j in TD:
    risyu_list.append(j.text)

# プログラム3-8-0　「授業科目一覧ページ3へ移動」    
page_third = browser.find_element_by_link_text('3')
page_third.click()
    
# プログラム3-8-1　「授業科目一覧ページのURLを取得」
cur_url = browser.current_url

# プログラム3-8-2　「授業科目一覧ページ（ブラウザで表示中）の全ソースを取得」
cur_source = browser.page_source

# プログラム3-8-3　「HTML解析と必要情報取得（と整理）」
soup = BeautifulSoup(cur_source, 'lxml')
TD = soup.find_all('td')

for j in TD:
    risyu_list.append(j.text)
    
# プログラム3-9　「検索画面へ戻る」
browser.back()
      
        
# プログラム4　「クローリングした内容を整える」

# プログラム4-1　「取得情報の整理」
risyu_list = list(map(lambda x : x.replace("\n",""),risyu_list))
risyu_list = [s.replace('        ○            ', '抽選科目') for s in risyu_list]
risyu_list = [s.replace('                -    ', '抽選なし') for s in risyu_list]
data = np.array(risyu_list).reshape(-1,7)
data = np.unique(data, axis=0)

# プログラム4-2　「不要科目削除」
# 大学院科目
data = data[data[:, 0] != '経営管理研究科']
data = data[data[:, 0] != '経済学研究科']
data = data[data[:, 0] != '法学研究科']
data = data[data[:, 0] != '社会学研究科']
data = data[data[:, 0] != '言語社会研究科']
data = data[data[:, 0] != '国際企業戦略研究科']
data = data[data[:, 0] != '国際・公共政策教育部']
data = data[data[:, 0] != '商学研究科']
data = data[data[:, 0] != '法科大学院']

# 全学共通教育科目ゼミ
data = data[data[:, 3] != '共通ゼミナール(3・4年)']
data = data[data[:, 3] != '共通ゼミナール（副）(3・4年)']
data = data[data[:, 3] != '共通ゼミナール(3年)']
data = data[data[:, 3] != '共通ゼミナール（副）(3年)']
data = data[data[:, 3] != '共通ゼミナール(4年)']
data = data[data[:, 3] != '共通ゼミナール（副）(4年)']

# 商学部ゼミ
data = data[data[:, 3] != '主ゼミナール(3年)']
data = data[data[:, 3] != '副ゼミナール(3年)']
data = data[data[:, 3] != '主ゼミナール(4年)']
data = data[data[:, 3] != '副ゼミナール(4年)']

# 経済学部ゼミ
data = data[data[:, 3] != 'ゼミナール（3年）']
data = data[data[:, 3] != 'ゼミナール（4年）']
data = data[data[:, 3] != '副ゼミナール(3年)']
data = data[data[:, 3] != '副ゼミナール(4年)']
data = data[data[:, 3] != 'ゼミナール（３年）']
data = data[data[:, 3] != 'ゼミナール（４年）']

# 法学部ゼミ
data = data[data[:, 3] != '主ゼミナール(3・4年)']
data = data[data[:, 3] != '副ゼミナール(3・4年)']
data = data[data[:, 3] != '主ゼミナール(3年)']
data = data[data[:, 3] != '副ゼミナール(3年)']
data = data[data[:, 3] != '主ゼミナール(4年)']
data = data[data[:, 3] != '副ゼミナール(4年)']

# 社会学部ゼミ
data = data[data[:, 3] != '主ゼミナール(3・4年)']
data = data[data[:, 3] != '副ゼミナール(3・4年)']
data = data[data[:, 3] != '主ゼミナール(3年)']
data = data[data[:, 3] != '副ゼミナール(3年)']
data = data[data[:, 3] != '主ゼミナール(4年)']
data = data[data[:, 3] != '副ゼミナール(4年)']

# PACE
data = data[data[:, 3] != 'PACEⅠ(再履修1)']
data = data[data[:, 3] != 'PACEⅡ(再履修1)']

for a in range(1, 63):
    data = data[data[:, 3] != 'PACEⅠ(' + str(a) + ')']

data = data[data[:, 3] != 'PACEⅠ(再履修2)']
data = data[data[:, 3] != 'PACEⅡ(再履修2)']

for b in range(1, 63):
    data = data[data[:, 3] != 'PACEⅡ(' + str(b) + ')']

# 卒業論文
data = data[data[:, 3] != '卒業論文(9月)']
data = data[data[:, 3] != '卒業論文(3月)']    

# 教育実習
data = data[data[:, 3] != '教育実習(中)']
data = data[data[:, 3] != '教育実習(高)'] 

# GEP
data = data[data[:, 1] != 'GEP（夏）']
data = data[data[:, 1] != 'GEP（秋）'] 
data = data[data[:, 1] != 'GEP（冬）']     


# プログラム5　「エクセルファイルへ出力する」    
    
# プログラム5-1　「エクセルを取得」
wb = px.Workbook()
ws = wb.active
 
# プログラム5-2　「エクセルのヘッダーの背景色を設定」
fill = PatternFill(patternType='solid', fgColor='e0e0e0', bgColor='e0e0e0')
 
# プログラム5-3　「エクセル1行目のヘッダーを出力」
headers = ['所属', '学期', '曜日・時限', '科目名','担当','教授言語','抽選対象']
for i, header in enumerate(headers):
    ws.cell(row=1, column=1+i, value=headers[i])
    ws.cell(row=1, column=1+i).fill = fill

# プログラム5-4　「エクセル2行目以降のデータを出力」
for y, row in enumerate(data):
    ws.cell(row= y+2, column= 1, value= y+1)
    for x, cell in enumerate(row):
        ws.cell(row= y+2, column= x+1, value=data[y][x])
        
# プログラム5-5　「セル幅の調整」
for col in ws.columns:
    max_length = 0
    column = col[0].column

    for cell in col:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))

    adjusted_width = (max_length + 2) * 1.5
    ws.column_dimensions[col[0].column_letter].width = adjusted_width

    
# プログラム5-6　「日付を取得（エクセルファイルに日付をつけてわかりやすくした！！）」
now = datetime.now()
hiduke = now.strftime('%Y-%m-%d')


# プログラム5-7　「エクセルファイルの保存」
filename = hiduke + '_' + 'シラバスクローリング' + '.xlsx'
wb.save(filename)

print("情報を取得し、エクセルファイルに保存しました。")
